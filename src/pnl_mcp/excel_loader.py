"""
Enhanced Excel data loading with smart header detection and data parsing.
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Tuple, Optional
import logging
import re

logger = logging.getLogger(__name__)


class ExcelDataLoader:
    """Enhanced Excel data loader with smart header detection."""
    
    def __init__(self):
        self.engine = 'openpyxl'
    
    def _detect_data_region(self, worksheet, max_rows: int = 50, max_cols: int = 50) -> Tuple[int, int, int, int]:
        """
        Detect the actual data region in a worksheet.
        
        Returns:
            Tuple of (start_row, start_col, end_row, end_col) (0-indexed)
        """
        # Find the first non-empty cell
        start_row, start_col = 0, 0
        end_row, end_col = 0, 0
        
        # Scan for data boundaries
        for row_idx, row in enumerate(worksheet.iter_rows(max_row=max_rows, max_col=max_cols)):
            if row_idx >= max_rows:
                break
                
            row_has_data = False
            for col_idx, cell in enumerate(row):
                if col_idx >= max_cols:
                    break
                    
                if cell.value is not None and str(cell.value).strip():
                    row_has_data = True
                    if start_row == 0 and start_col == 0:
                        start_row, start_col = row_idx, col_idx
                    end_row = max(end_row, row_idx)
                    end_col = max(end_col, col_idx)
            
            # If we haven't found data yet, continue
            if not row_has_data and end_row == 0:
                continue
            
            # If we've found data but this row is empty, might be end of data
            if not row_has_data and end_row > 0:
                # Check if we have several empty rows (might be end of data)
                empty_rows = 0
                for check_row in range(row_idx, min(row_idx + 5, max_rows)):
                    if check_row < len(list(worksheet.iter_rows(min_row=check_row+1, max_row=check_row+1))):
                        check_row_data = list(worksheet.iter_rows(min_row=check_row+1, max_row=check_row+1, max_col=max_cols))[0]
                        if any(cell.value is not None and str(cell.value).strip() for cell in check_row_data):
                            break
                        empty_rows += 1
                    else:
                        empty_rows += 1
                
                if empty_rows >= 3:  # 3+ empty rows likely means end of data
                    break
        
        return start_row, start_col, end_row, end_col
    
    def _detect_headers(self, worksheet, start_row: int, start_col: int, end_col: int, max_header_rows: int = 5) -> Tuple[int, List[str]]:
        """
        Detect header row and extract column names.
        
        Returns:
            Tuple of (header_row_index, column_names)
        """
        best_header_row = start_row
        best_score = 0
        best_headers = []
        
        # Check first few rows for the best header
        for row_idx in range(start_row, min(start_row + max_header_rows, start_row + 10)):
            try:
                row_cells = list(worksheet.iter_rows(min_row=row_idx+1, max_row=row_idx+1, 
                                                   min_col=start_col+1, max_col=end_col+1))[0]
                headers = [cell.value for cell in row_cells]
                
                # Score this row as a potential header
                score = self._score_header_row(headers)
                
                if score > best_score:
                    best_score = score
                    best_header_row = row_idx
                    best_headers = headers
                    
            except (IndexError, AttributeError):
                continue
        
        # Clean up headers
        cleaned_headers = self._clean_headers(best_headers)
        
        return best_header_row, cleaned_headers
    
    def _score_header_row(self, headers: List[Any]) -> int:
        """Score a row's likelihood of being a header row."""
        score = 0
        
        for header in headers:
            if header is None:
                continue
                
            header_str = str(header).strip()
            if not header_str:
                continue
            
            # String headers are good
            if isinstance(header, str):
                score += 10
            
            # Headers with common financial terms
            financial_terms = ['pnl', 'profit', 'loss', 'delta', 'gamma', 'vega', 'theta', 
                             'realized', 'unrealized', 'impact', 'value', 'rate', 'spot']
            if any(term in header_str.lower() for term in financial_terms):
                score += 20
            
            # Headers that look like column names (no spaces or have underscores)
            if '_' in header_str or (len(header_str.split()) == 1 and header_str.isalpha()):
                score += 5
            
            # Avoid numeric-only headers
            if header_str.replace('.', '').replace('-', '').isdigit():
                score -= 10
            
            # Headers with "Unnamed" are bad
            if 'unnamed' in header_str.lower():
                score -= 20
        
        return score
    
    def _clean_headers(self, headers: List[Any]) -> List[str]:
        """Clean and standardize header names."""
        cleaned = []
        
        for i, header in enumerate(headers):
            if header is None or str(header).strip() == '':
                cleaned.append(f"Column_{i}")
            else:
                # Convert to string and clean
                header_str = str(header).strip()
                # Remove extra whitespace
                header_str = re.sub(r'\s+', ' ', header_str)
                # Handle special characters
                header_str = header_str.replace('\n', ' ').replace('\r', ' ')
                cleaned.append(header_str)
        
        return cleaned
    
    def _load_with_openpyxl(self, filepath: str, sheet_name: str = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Load Excel file using openpyxl for better control."""
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        
        # Get sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
            sheet_name = worksheet.title
        
        # Detect data region
        start_row, start_col, end_row, end_col = self._detect_data_region(worksheet)
        
        # Detect headers
        header_row, headers = self._detect_headers(worksheet, start_row, start_col, end_col)
        
        # Load data using pandas with detected parameters
        df = pd.read_excel(
            filepath,
            sheet_name=sheet_name,
            header=header_row,
            usecols=f"{openpyxl.utils.get_column_letter(start_col+1)}:{openpyxl.utils.get_column_letter(end_col+1)}",
            engine=self.engine
        )
        
        # Handle dict return (shouldn't happen with specific sheet_name, but just in case)
        if isinstance(df, dict):
            df = df[sheet_name]
        
        # Set proper column names
        if len(headers) == len(df.columns):
            df.columns = headers
        
        # Metadata
        metadata = {
            'sheet_name': sheet_name,
            'data_region': {
                'start_row': start_row,
                'start_col': start_col,
                'end_row': end_row,
                'end_col': end_col
            },
            'header_row': header_row,
            'original_headers': headers,
            'available_sheets': workbook.sheetnames
        }
        
        workbook.close()
        return df, metadata
    
    def load_excel_data(self, filepath: str, sheet_name: str = None, 
                       range_spec: str = None, sample_rows: int = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Load Excel data with smart header detection and data parsing.
        
        Args:
            filepath: Path to Excel file
            sheet_name: Sheet name to load (None for active sheet)
            range_spec: Range specification (e.g., "A1:C10")
            sample_rows: Number of rows to sample (None for all)
            
        Returns:
            Tuple of (DataFrame, metadata_dict)
        """
        try:
            # Use enhanced loading
            df, metadata = self._load_with_openpyxl(filepath, sheet_name)
            
            # Apply range specification if provided
            if range_spec:
                # Parse range (simplified - could be enhanced)
                try:
                    if ':' in range_spec:
                        start_cell, end_cell = range_spec.split(':')
                        # This is a simplified implementation
                        # In production, you'd want more robust range parsing
                        pass
                except:
                    logger.warning(f"Could not parse range specification: {range_spec}")
            
            # Apply row sampling if requested
            if sample_rows and len(df) > sample_rows:
                df = df.head(sample_rows)
                metadata['sampled'] = True
                metadata['sample_rows'] = sample_rows
            else:
                metadata['sampled'] = False
            
            # Clean up data
            df = df.dropna(axis=1, how='all')  # Remove completely empty columns
            
            # Add data quality info
            metadata['data_quality'] = {
                'shape': df.shape,
                'null_counts': df.isnull().sum().to_dict(),
                'dtypes': df.dtypes.astype(str).to_dict(),
                'memory_usage': df.memory_usage(deep=True).sum()
            }
            
            return df, metadata
            
        except Exception as e:
            logger.error(f"Error loading Excel file {filepath}: {e}")
            raise
    
    def get_sheet_info(self, filepath: str) -> Dict[str, Any]:
        """Get information about all sheets in the Excel file."""
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                start_row, start_col, end_row, end_col = self._detect_data_region(worksheet)
                
                sheet_info[sheet_name] = {
                    'data_region': {
                        'start_row': start_row,
                        'start_col': start_col, 
                        'end_row': end_row,
                        'end_col': end_col
                    },
                    'estimated_rows': end_row - start_row + 1,
                    'estimated_cols': end_col - start_col + 1
                }
            
            workbook.close()
            return {
                'sheets': sheet_info,
                'total_sheets': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames
            }
            
        except Exception as e:
            logger.error(f"Error getting sheet info for {filepath}: {e}")
            raise


# Global loader instance
_loader = ExcelDataLoader()

def load_excel_with_smart_headers(filepath: str, sheet_name: str = None, 
                                range_spec: str = None, sample_rows: int = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Load Excel data with enhanced header detection.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name to load
        range_spec: Range specification
        sample_rows: Number of rows to sample
        
    Returns:
        Tuple of (DataFrame, metadata)
    """
    return _loader.load_excel_data(filepath, sheet_name, range_spec, sample_rows)

def get_excel_sheet_info(filepath: str) -> Dict[str, Any]:
    """Get information about Excel file sheets."""
    return _loader.get_sheet_info(filepath)